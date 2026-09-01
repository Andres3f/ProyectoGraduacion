from datetime import date, datetime, time
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.route import Route
from app.models.route_stop import RouteStop
from app.models.order import Order
from app.models.user import User, RoleEnum
from app.auth.dependencies import require_role
from app.services.metrics import compare_before_after, estimate_savings

router = APIRouter(prefix="/api/dashboard", tags=["Dashboard / KPIs"])


def _range_start_end(date_from: date, date_to: date):
    """Convierte fechas de un día a intervalos datetime (inicio/fin de día)."""
    start = datetime.combine(date_from, time.min)
    end = datetime.combine(date_to, time.max)
    return start, end


def _build_report(db: Session, date_from: date, date_to: date) -> dict:
    """Recopila KPIs + detalle de rutas del rango para reportes/exportación."""
    start, end = _range_start_end(date_from, date_to)

    routes = db.query(Route).filter(Route.created_at.between(start, end)).all()
    total_km = sum(r.total_distance_km or 0 for r in routes)

    stops = (
        db.query(RouteStop)
        .join(Route)
        .filter(Route.created_at.between(start, end))
        .all()
    )
    delivered = [s for s in stops if s.status == "entregado"]
    on_time = [
        s for s in delivered
        if s.delivered_at and s.eta and s.delivered_at <= s.eta
    ]

    kpis = {
        "total_distance_km": round(total_km, 2),
        "total_routes": len(routes),
        "delivery_rate_pct": round(len(delivered) / len(stops) * 100, 1) if stops else 0,
        "on_time_rate_pct": round(len(on_time) / len(delivered) * 100, 1) if delivered else 0,
        "avg_km_per_route": round(total_km / len(routes), 2) if routes else 0,
    }

    # PG-22: ahorro estimado de combustible y de costo operativo. La distancia
    # ahorrada es la suma (por ruta) de distancia antes - después.
    routes_rows = [_route_dashboard_row(db, r) for r in routes]
    total_km_saved = sum(
        max(0, row.get("distance_before_km", 0) - row.get("distance_after_km", 0))
        for row in routes_rows
    )
    savings = estimate_savings(total_km_saved)
    kpis.update(savings)

    # PG-23: distribución de entregas (entregado / fallido / pendiente).
    stop_status_counts = {
        "entregado": len([s for s in stops if s.status == "entregado"]),
        "fallido": len([s for s in stops if s.status == "fallido"]),
        "pendiente": len([s for s in stops if s.status == "pendiente"]),
    }
    kpis["delivery_distribution"] = stop_status_counts

    return {
        "kpis": kpis,
        "routes": routes_rows,
    }


def _route_dashboard_row(db: Session, r: Route) -> dict:
    """Convierte una ruta en una fila de reporte con distancia antes/después."""
    # Distancia "antes": recorrido naive en el orden original de los pedidos.
    # Distancia "después": la distancia optimizada ya persistida en la ruta.
    optimized = [
        {"lat": s.order.latitude, "lng": s.order.longitude}
        for s in (r.stops or [])
        if s.order
    ]
    row = {
        "name": r.name or f"Ruta #{r.id}",
        "id": r.id,
        "status": r.status.value,
        "vehicle_id": r.vehicle_id,
        "driver": db.get(User, r.driver_id).full_name if r.driver_id else "—",
        "distance_km": r.total_distance_km or 0,
        "stops": len(r.stops or []),
        "optimized_at": r.optimized_at,
        "created_at": r.created_at,
    }
    if optimized:
        orders_naive = (
            db.query(Order)
            .join(RouteStop)
            .filter(RouteStop.route_id == r.id)
            .order_by(Order.id)
            .all()
        )
        metrics = compare_before_after(orders_naive, optimized)
        row["distance_before_km"] = metrics.get("distance_before_km", 0)
        row["distance_after_km"] = metrics.get("distance_after_km", 0)
    else:
        row["distance_before_km"] = 0
        row["distance_after_km"] = r.total_distance_km or 0
    return row


@router.get("/kpis")
def get_kpis(
    date_from: date = Query(..., description="Fecha inicial (YYYY-MM-DD)"),
    date_to: date = Query(..., description="Fecha final (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_role([RoleEnum.gerente, RoleEnum.admin])),
):
    """Indicadores clave del negocio en un rango de fechas (OPT-22)."""
    return _build_report(db, date_from, date_to)["kpis"]


@router.get("/kpis/timeseries")
def get_kpis_timeseries(
    date_from: date = Query(..., description="Fecha inicial (YYYY-MM-DD)"),
    date_to: date = Query(..., description="Fecha final (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    _: User = Depends(require_role([RoleEnum.gerente, RoleEnum.admin])),
):
    """Devuelve un punto por día con % de reducción y km ahorrados (PG-23).

    Punto de datos para el gráfico de línea de evolución de la reducción.
    """
    routes_rows = _build_report(db, date_from, date_to)["routes"]

    per_day: dict[date, dict] = {}
    for r in routes_rows:
        day = r["created_at"].date() if r.get("created_at") else None
        if day is None:
            continue
        entry = per_day.setdefault(
            day, {"before": 0.0, "after": 0.0, "km_saved": 0.0}
        )
        before = r.get("distance_before_km", 0)
        after = r.get("distance_after_km", 0)
        entry["before"] += before
        entry["after"] += after
        entry["km_saved"] += max(0, before - after)

    series = []
    for day in sorted(per_day):
        e = per_day[day]
        reduction = (
            round((1 - e["after"] / e["before"]) * 100, 1) if e["before"] else 0
        )
        series.append({
            "date": day.isoformat(),
            "reduction_percentage": reduction,
            "km_saved": round(e["km_saved"], 2),
            "routes": sum(1 for r in routes_rows if r.get("created_at") and r["created_at"].date() == day),
        })

    return {"dates": [s["date"] for s in series], "series": series}


@router.get("/export")
def export_report(
    format: str = Query("xlsx", description="xlsx | pdf"),
    date_from: date = Query(..., description="Fecha inicial (YYYY-MM-DD)"),
    date_to: date = Query(..., description="Fecha final (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([RoleEnum.gerente, RoleEnum.admin])),
):
    """Descarga un reporte del rango (KPIs + tabla de rutas) (OPT-23)."""
    data = _build_report(db, date_from, date_to)
    kpis = data["kpis"]
    routes = data["routes"]

    if format == "xlsx":
        content, media_type, filename = _build_xlsx(kpis, routes, date_from, date_to)
    elif format == "pdf":
        content, media_type, filename = _build_pdf(kpis, routes, date_from, date_to)
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado (usa xlsx o pdf)")

    from app.services.audit import log_action

    log_action(
        db, current_user.id, "exportar_reporte",
        entidad="report", detalle={"format": format, "date_from": date_from.isoformat(), "date_to": date_to.isoformat()},
    )
    db.commit()

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_xlsx(kpis, routes, date_from, date_to):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(["Reporte Optirutas Jalapa"])
    ws.append([f"Rango: {date_from} a {date_to}"])
    ws.append([])
    ws.append(["KPIs"])
    ws.append(["Distancia total (km)", kpis["total_distance_km"]])
    ws.append(["Rutas totales", kpis["total_routes"]])
    ws.append(["Tasa de entrega (%)", kpis["delivery_rate_pct"]])
    ws.append(["A tiempo (%)", kpis["on_time_rate_pct"]])
    ws.append(["km promedio por ruta", kpis["avg_km_per_route"]])
    ws.append([])

    ws.append(["Rutas"])
    header = ["Ruta", "Estado", "Vehículo", "Conductor", "Distancia (km)", "Paradas"]
    ws.append(header)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for r in routes:
        ws.append([
            r["name"], r["status"], r["vehicle_id"], r["driver"],
            r["distance_km"], r["stops"],
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"optirutas_{date_from}_{date_to}.xlsx"


def _build_pdf(kpis, routes, date_from, date_to):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte Optirutas Jalapa", styles["Title"]))
    story.append(Paragraph(f"Rango: {date_from} a {date_to}", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("KPIs", styles["Heading2"]))
    kpi_rows = [
        ["Distancia total (km)", kpis["total_distance_km"]],
        ["Rutas totales", kpis["total_routes"]],
        ["Tasa de entrega (%)", kpis["delivery_rate_pct"]],
        ["A tiempo (%)", kpis["on_time_rate_pct"]],
        ["km promedio por ruta", kpis["avg_km_per_route"]],
    ]
    kpi_table = Table(kpi_rows)
    kpi_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Rutas", styles["Heading2"]))
    header = ["Ruta", "Estado", "Vehículo", "Conductor", "Km", "Paradas"]
    rows = [header] + [
        [r["name"], r["status"], r["vehicle_id"], r["driver"],
         r["distance_km"], r["stops"]]
        for r in routes
    ]
    table = Table(rows, colWidths=[120, 70, 55, 90, 45, 45])
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue(), "application/pdf", f"optirutas_{date_from}_{date_to}.pdf"
